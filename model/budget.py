import json
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class OverBudgetError(Exception):
    pass

class Category:
    def __init__(self, name, percentage):
        self.name = name
        self.percentage = percentage  # In percentage mode, the imported (or computed) percentage.
        self.amount = 0               # In amount mode, the imported (or computed) amount.
        self.amount_override = None   # If user edits the amount while locked.
        self.lock_type = 0            # 0 = Unlocked, 1 = Lock Amount, 2 = Lock Percentage

    def to_dict(self):
        return {
            "name": self.name,
            "percentage": self.percentage,
            "amount": self.amount,
            "amount_override": self.amount_override,
            "lock_type": self.lock_type
        }

    @classmethod
    def from_dict(cls, data):
        cat = cls(data["name"], data["percentage"])
        cat.amount = data.get("amount", 0)
        cat.amount_override = data.get("amount_override")
        cat.lock_type = data.get("lock_type", 0)
        return cat

class FeeItem:
    def __init__(self, name, fee_type, value):
        self.name = name
        self.fee_type = fee_type.lower()  # "percentage" or "fixed"
        self.value = value                # For percentage fees: whole-number percent (e.g., 5 for 5%); for fixed fees: the amount.
        self.computed_amount = 0          # For percentage fees, computed from subtotal.

    def to_dict(self):
        return {
            "name": self.name,
            "fee_type": self.fee_type,
            "value": self.value
        }

    @classmethod
    def from_dict(cls, data):
        return cls(data["name"], data.get("fee_type", "percentage"), data.get("value", 0.0))

class BudgetModel:
    def __init__(self, grand_total, admin_pct=5.0, contingency_pct=10.0, categories=None):
        self.grand_total = grand_total
        self.admin_pct = admin_pct
        self.contingency_pct = contingency_pct
        self.fees = []  # List of FeeItem objects (if any)
        if categories is None:
            categories = []
        self.categories = categories
        # Default groups – used only if a manual budget is created.
        self.groups = [
            ("TOTAL SCRIPT AND DEVELOPMENT", list(range(0, 4))),
            ("TOTAL PRODUCTION COSTS", list(range(4, 14))),
            ("TOTAL POST PRODUCTION", list(range(14, 18))),
            ("TOTAL OTHER COSTS", list(range(18, 21)))
        ]
        self.over_budget = False
        self.over_budget_rows = []
        self.import_mode = "amount"  # "amount" or "percentage"
        self.computed_grand_total = grand_total
        # New flag: if True, keep the imported category amounts (do not recalc them).
        self.keep_category_amounts = False
        self.recalc()

    def recalc(self):
        # Calculate subtotal.
        if not self.fees:
            self.subtotal = self.grand_total
        else:
            if self.import_mode == "amount":
                self.subtotal = sum(cat.amount for cat in self.categories)
            else:
                fixed_fee_total = sum(fee.value for fee in self.fees if fee.fee_type == "fixed")
                total_fee_pct = sum(fee.value for fee in self.fees if fee.fee_type == "percentage")
                if (1 + total_fee_pct/100) != 0:
                    self.subtotal = round((self.grand_total - fixed_fee_total) / (1 + total_fee_pct/100))
                else:
                    self.subtotal = 0
        logger.debug("Recalc: grand_total=%s, subtotal=%s", self.grand_total, self.subtotal)
        # Recalculate categories.
        if self.import_mode == "amount" and self.keep_category_amounts:
            # Preserve the imported amounts; just update percentages.
            if self.subtotal:
                for cat in self.categories:
                    cat.percentage = (cat.amount / self.subtotal * 100)
            else:
                for cat in self.categories:
                    cat.percentage = 0
        else:
            # Normal percentage-mode recalculation.
            if self.import_mode == "amount":
                # In amount mode, update percentages based on amounts.
                for cat in self.categories:
                    if self.subtotal:
                        cat.percentage = (cat.amount / self.subtotal * 100)
                    else:
                        cat.percentage = 0
            else:
                fixed_percentage_total = 0.0
                for cat in self.categories:
                    if cat.lock_type == 1:  # Lock Amount
                        fixed_amt = cat.amount_override if cat.amount_override is not None else cat.amount
                        cat.amount = fixed_amt
                        cat.percentage = (fixed_amt / self.subtotal * 100) if self.subtotal else 0
                        fixed_percentage_total += cat.percentage
                    elif cat.lock_type == 2:  # Lock Percentage
                        cat.amount = round(self.subtotal * (cat.percentage / 100))
                        fixed_percentage_total += cat.percentage
                unlocked = [cat for cat in self.categories if cat.lock_type == 0]
                unlocked_count = len(unlocked)
                available_pct = 100 - fixed_percentage_total
                if unlocked_count > 0:
                    sum_desired = sum(cat.percentage for cat in unlocked)
                    if sum_desired <= 0:
                        for cat in unlocked:
                            cat.percentage = available_pct / unlocked_count
                            cat.amount = round(self.subtotal * (cat.percentage / 100))
                    else:
                        for cat in unlocked:
                            new_pct = (cat.percentage / sum_desired) * available_pct
                            cat.percentage = new_pct
                            cat.amount = round(self.subtotal * (new_pct / 100))
        # Recalculate fee amounts.
        for fee in self.fees:
            if fee.fee_type == "percentage":
                fee.computed_amount = round(self.subtotal * (fee.value / 100))
            else:
                fee.computed_amount = fee.value
        if self.fees:
            total_fee = sum(fee.computed_amount for fee in self.fees)
            self.computed_grand_total = self.subtotal + total_fee
        else:
            self.computed_grand_total = self.subtotal
        self.check_over_budget()

    def check_over_budget(self):
        tolerance = 1
        fixed_total = sum(cat.amount for cat in self.categories if cat.lock_type in [1, 2])
        if fixed_total > self.subtotal + tolerance:
            self.over_budget = True
            self.over_budget_rows = [i for i, cat in enumerate(self.categories) if cat.lock_type in [1, 2]]
        else:
            self.over_budget = False
            self.over_budget_rows = []

    def update_category_percentage(self, index, new_percentage):
        new_percentage = max(new_percentage, 0)
        self.categories[index].percentage = new_percentage
        if self.categories[index].lock_type != 2:
            self.categories[index].amount_override = None
        # Editing a category clears the "keep" flag.
        self.keep_category_amounts = False
        logger.debug("update_category_percentage: index=%s, new_percentage=%s", index, new_percentage)
        self.recalc()

    def update_category_amount(self, index, new_amount):
        new_amount = max(new_amount, 0)
        self.categories[index].amount_override = new_amount
        if self.subtotal:
            self.categories[index].percentage = new_amount / self.subtotal * 100
        else:
            self.categories[index].percentage = 0
        # Editing a category clears the "keep" flag.
        self.keep_category_amounts = False
        logger.debug("update_category_amount: index=%s, new_amount=%s", index, new_amount)
        self.recalc()

    def update_lock_type(self, index, lock_type):
        self.categories[index].lock_type = lock_type
        if lock_type == 0:
            self.categories[index].amount_override = None
        self.keep_category_amounts = False
        logger.debug("update_lock_type: index=%s, new_lock_type=%s", index, lock_type)
        self.recalc()

    def lock_all(self, lock_type):
        for i in range(len(self.categories)):
            self.categories[i].lock_type = lock_type
        self.recalc()

    def unlock_all(self):
        for i in range(len(self.categories)):
            self.categories[i].lock_type = 0
            self.categories[i].amount_override = None
        self.recalc()

    def set_grand_total(self, new_total):
        self.grand_total = max(new_total, 0)
        self.recalc()

    def set_admin_pct(self, new_pct):
        self.admin_pct = max(new_pct, 0)
        self.recalc()

    def set_contingency_pct(self, new_pct):
        self.contingency_pct = max(new_pct, 0)
        self.recalc()

    def get_group_total(self, indices):
        return sum(self.categories[i].amount for i in indices)

    def get_group_percentage(self, indices):
        group_amt = self.get_group_total(indices)
        return (group_amt / self.subtotal * 100) if self.subtotal else 0

    def get_table_data(self):
        data = []
        group_mapping = {}
        for group in self.groups:
            label, indices = group
            member_rows = []
            for idx in indices:
                if idx < len(self.categories):
                    cat = self.categories[idx]
                    row = {
                        "row_type": "category",
                        "description": cat.name,
                        "amount": cat.amount,
                        "percentage": cat.percentage,
                        "lock_type": cat.lock_type,
                        "cat_index": idx
                    }
                    member_rows.append(len(data))
                    data.append(row)
            total_row = {
                "row_type": "group_total",
                "description": label,
                "amount": self.get_group_total(indices),
                "percentage": self.get_group_percentage(indices),
                "group_label": label
            }
            group_mapping[label] = {"member_rows": member_rows, "total_row": len(data)}
            data.append(total_row)
        data.append({
            "row_type": "subtotal",
            "description": "SUBTOTAL",
            "amount": self.subtotal,
            "percentage": None
        })
        for i, fee in enumerate(self.fees):
            if fee.fee_type == "percentage":
                row = {
                    "row_type": "fee",
                    "description": fee.name,
                    "amount": fee.computed_amount,
                    "percentage": fee.value,
                    "fee_index": i
                }
            else:
                fee_pct = (fee.value / self.subtotal * 100) if self.subtotal > 0 else 0
                row = {
                    "row_type": "fee",
                    "description": fee.name,
                    "amount": fee.value,
                    "percentage": fee_pct,
                    "fee_index": i
                }
            row["editable"] = True
            data.append(row)
        data.append({
            "row_type": "grand_total",
            "description": "GRAND TOTAL",
            "amount": self.computed_grand_total,
            "percentage": None
        })
        return data, group_mapping

    def to_dict(self):
        groups_data = []
        for group in self.groups:
            label, indices = group
            items = [self.categories[i].to_dict() for i in indices if i < len(self.categories)]
            groups_data.append({"name": label, "items": items})
        return {
            "grand_total": self.grand_total,
            "admin_pct": self.admin_pct,
            "contingency_pct": self.contingency_pct,
            "categories": [cat.to_dict() for cat in self.categories],
            "groups": groups_data,
            "fees": [fee.to_dict() for fee in self.fees],
            "import_mode": self.import_mode
        }

    def from_dict(self, data):
        self.grand_total = data.get("grand_total", 0)
        self.admin_pct = data.get("admin_pct", 0)
        self.contingency_pct = data.get("contingency_pct", 0)
        self.categories = [Category.from_dict(c) for c in data.get("categories", [])]
        self.fees = [FeeItem.from_dict(f) for f in data.get("fees", [])]
        self.import_mode = data.get("import_mode", "amount")
        self.recalc()

    def save_to_file(self, filename):
        with open(filename, "w") as f:
            json.dump(self.to_dict(), f, indent=4)

    def load_from_file(self, filename):
        with open(filename, "r") as f:
            data = json.load(f)
        self.from_dict(data)

    def import_from_excel(self, filename):
        wb = load_workbook(filename=filename, data_only=True)
        ws = wb.active
        groups_dict = {}
        categories_list = []
        group_order = []
        self.fees = []  # clear current fees
        mode_detected = None  # "amount" or "percentage"
        # Assume first row is header.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            group = str(row[0]).strip()
            cat_desc = str(row[1]).strip() if row[1] is not None else ""
            amt_cell = str(row[2]).strip() if row[2] is not None else ""
            pct_cell = str(row[3]).strip() if row[3] is not None else ""
            try:
                amount = float(amt_cell.replace(" ", "").replace(",", "."))
            except Exception:
                amount = 0.0
            try:
                percentage = float(pct_cell.replace(" ", "").replace(",", "."))
            except Exception:
                percentage = 0.0
            if group.upper().startswith("FEES"):
                fee_type = "percentage" if amount == 0 else "fixed"
                if fee_type == "percentage" and percentage < 1:
                    percentage *= 100
                self.fees.append(FeeItem(cat_desc, fee_type, percentage if fee_type=="percentage" else amount))
            else:
                if mode_detected is None:
                    mode_detected = "amount" if amount > 0 else "percentage"
                else:
                    current_mode = "amount" if amount > 0 else "percentage"
                    if current_mode != mode_detected:
                        raise ValueError("Mixed mode detected in Excel import. Please use either amounts or percentages exclusively.")
                cat = Category(cat_desc, percentage)
                cat.amount = amount
                cat_index = len(categories_list)
                categories_list.append(cat)
                if group not in groups_dict:
                    groups_dict[group] = []
                    group_order.append(group)
                groups_dict[group].append(cat_index)
        self.categories = categories_list
        self.groups = []
        for group in group_order:
            self.groups.append((group, groups_dict[group]))
        # Conversion: if mode_detected is "amount", then preserve category amounts.
        if mode_detected == "amount":
            cat_total = sum(cat.amount for cat in self.categories)
            if cat_total > 0:
                for cat in self.categories:
                    cat.percentage = (cat.amount / cat_total * 100)
            # Convert fee fixed amounts to percentages relative to cat_total.
            for fee in self.fees:
                if fee.fee_type == "fixed":
                    fee.value = (fee.value / cat_total) * 100
                    fee.fee_type = "percentage"
            self.grand_total = cat_total + sum(fee.value/100 * cat_total for fee in self.fees if fee.fee_type=="percentage")
            # Set flag to preserve category amounts.
            self.keep_category_amounts = True
            self.import_mode = "percentage"
        else:
            self.import_mode = "percentage"
        self.recalc()
